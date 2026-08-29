"""Investment export parsers (Phase 1) — Nordnet CSV + Nordea Omistukset.xlsx.

Parses broker portfolio exports into a normalized hierarchy:

    Broker (main)  ->  Account (sub, if available)  ->  Holding (leaf)
                                                         name, units (pcs),
                                                         value_eur, development

Pure module (no Flask). Phase 2 maps the result onto accounts / account_balances
/ holdings and writes today's snapshot; Phase 3 renders the grouped UI.

Real-format notes (verified against the actual exports):
- Nordnet CSVs: UTF-16 LE, TAB-delimited, Finnish headers, decimal comma. EUR
  value = "Arvo EUR"; units = "Määrä"; return = "Tuotto"/"Tuotto, %" + "Tuotto, EUR".
  No date inside the file -> taken from the filename (DD.M.YYYY).
- Nordea Omistukset.xlsx: sheet "Holdings"; row 0 = export timestamp; row 1 =
  headers; Type = Custody (investment) | CashAccount (deposit). EUR value =
  "Value in base currency" (custody) / "Value on account level" (cash).

Run `python3 investment_import.py` to verify parsing against the files in
~/Downloads (read-only, no DB writes).
"""

import csv
import io
import os
import re
from dataclasses import dataclass, field

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


class ImportError_(Exception):
    """Unreadable or unrecognized export file."""


@dataclass
class Holding:
    name: str
    value_eur: float
    units: float | None = None
    isin: str | None = None
    return_pct: float | None = None
    return_eur: float | None = None
    currency: str | None = None


@dataclass
class Account:
    broker: str                       # main category
    label: str                        # sub category (account name / portfolio no.)
    external_id: str                  # stable map key, e.g. "csv:nordnet:18318444"
    kind: str                         # "investment" | "cash"
    holdings: list[Holding] = field(default_factory=list)
    cash_value: float | None = None   # for kind == "cash"

    @property
    def total_eur(self) -> float:
        if self.kind == "cash":
            return self.cash_value or 0.0
        return round(sum(h.value_eur for h in self.holdings), 2)


@dataclass
class ParsedFile:
    source: str                       # "nordnet_stocks" | "nordnet_funds" | "nordea_xlsx"
    as_of: str | None                 # YYYY-MM-DD
    accounts: list[Account] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ── number / header helpers ──────────────────────────────────────────

def parse_num(s) -> float | None:
    """Parse a number that may use a comma or period decimal, with nbsp/space
    thousands separators. Returns None for blanks/non-numeric."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    t = str(s).replace("\xa0", "").replace(" ", "").strip()
    if t in ("", "-", "nan", "None"):
        return None
    # If both separators present, the last one is the decimal mark.
    if "," in t and "." in t:
        if t.rfind(",") > t.rfind("."):
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
    else:
        t = t.replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return None


def _norm(h) -> str:
    return str(h or "").replace("\xa0", " ").strip().lower()


def _col_finder(header):
    """Return f(*candidates) -> index, matching normalized header names
    (exact first, then 'startswith', then 'contains')."""
    norm = [_norm(h) for h in header]

    def find(*cands, contains=False):
        cands = [c.lower() for c in cands]
        for c in cands:
            if c in norm:
                return norm.index(c)
        for c in cands:
            for i, h in enumerate(norm):
                if h.startswith(c):
                    return i
        if contains:
            for c in cands:
                for i, h in enumerate(norm):
                    if c in h:
                        return i
        return None
    return find


def _date_from_filename(filename: str) -> str | None:
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", filename or "")
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None


# ── Nordnet CSV (stocks / funds) ─────────────────────────────────────

def parse_nordnet_csv(filename: str, raw: bytes) -> ParsedFile:
    try:
        text = raw.decode("utf-16")
    except UnicodeError:
        text = raw.decode("utf-8-sig", errors="replace")
    first = text.splitlines()[0] if text else ""
    delim = "\t" if "\t" in first else (";" if ";" in first else ",")
    rows = [r for r in csv.reader(io.StringIO(text), delimiter=delim) if any(c.strip() for c in r)]
    if not rows:
        raise ImportError_(f"{filename}: empty CSV")
    header = rows[0]
    is_funds = _norm(header[0]) == "rahastot"
    find = _col_finder(header)
    i_name = 0
    i_units = find("määrä", "antal", "amount")
    i_value = find("arvo eur", "värde eur", "market value eur")
    i_ret_pct = find("tuotto, %", "tuotto %", "tuotto", "avkastning")
    i_ret_eur = find("tuotto, eur", "tuotto eur", "avkastning eur")
    i_cur = find("valuutta", "valuta", "currency")
    if i_value is None:
        raise ImportError_(f"{filename}: no 'Arvo EUR' column found")

    salkku = re.search(r"salkku\w*\s*(\d+)", filename or "", re.I)
    acct_label = f"Nordnet {salkku.group(1)}" if salkku else "Nordnet"
    ext = f"csv:nordnet:{salkku.group(1)}" if salkku else "csv:nordnet"

    acct = Account(broker="Nordnet", label=acct_label, external_id=ext, kind="investment")
    warnings = []
    for r in rows[1:]:
        if i_name >= len(r):
            continue
        name = (r[i_name] or "").strip()
        val = parse_num(r[i_value]) if i_value < len(r) else None
        if not name or val is None:
            continue
        # Nordnet "Tuotto" on funds is a plain % (e.g. "2" = 2.00). Keep as-is.
        ret_pct = parse_num(r[i_ret_pct]) if i_ret_pct is not None and i_ret_pct < len(r) else None
        ret_eur = parse_num(r[i_ret_eur]) if i_ret_eur is not None and i_ret_eur < len(r) else None
        acct.holdings.append(Holding(
            name=name, value_eur=val,
            units=parse_num(r[i_units]) if i_units is not None and i_units < len(r) else None,
            return_pct=ret_pct, return_eur=ret_eur,
            currency=(r[i_cur].strip() if i_cur is not None and i_cur < len(r) else None),
        ))
    if not acct.holdings:
        warnings.append(f"{filename}: no holdings parsed")
    return ParsedFile(
        source="nordnet_funds" if is_funds else "nordnet_stocks",
        as_of=_date_from_filename(filename),
        accounts=[acct], warnings=warnings,
    )


# ── Nordea Omistukset.xlsx ───────────────────────────────────────────

def parse_nordea_xlsx(filename: str, raw: bytes) -> ParsedFile:
    if openpyxl is None:
        raise ImportError_("openpyxl is required to read .xlsx files")
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["Holdings"] if "Holdings" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 3:
        raise ImportError_(f"{filename}: too few rows")

    # Row 0 = export timestamp; row 1 = headers.
    as_of = None
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", str(rows[0][0] or ""))
    if m:
        d, mo, y = m.groups()
        as_of = f"{y}-{int(mo):02d}-{int(d):02d}"
    header = rows[1]
    find = _col_finder(header)
    i_type = find("type")
    i_acct = find("account")               # exact 'account' (not 'accountkey')
    i_isin = find("isin")
    i_cur = find("currency")
    i_name = find("name")
    i_units = find("holdings")
    i_val_base = find("value in base currency")
    i_val_acct = find("value on account level")
    i_ret_eur = find("value change on account level")
    i_purch = find("purchase value of currency", "purchase value", contains=True)
    if i_type is None or i_acct is None:
        raise ImportError_(f"{filename}: unexpected Holdings layout")

    def cell(r, i):
        return r[i] if i is not None and i < len(r) else None

    invest_accts: dict[str, Account] = {}
    cash_accts: list[Account] = []
    warnings = []
    for r in rows[2:]:
        rtype = _norm(cell(r, i_type))
        acct_str = str(cell(r, i_acct) or "").strip()
        if not rtype:
            continue
        if rtype == "cashaccount":
            val = parse_num(cell(r, i_val_acct)) or parse_num(cell(r, i_val_base))
            if val is None:
                continue
            iban = re.search(r"\b([A-Z]{2}\d{2}[\d ]{6,})", acct_str)
            ext = "csv:nordea:" + (re.sub(r"\s", "", iban.group(1)) if iban else acct_str)
            cash_accts.append(Account(
                broker="Nordea", label=acct_str or "Nordea account",
                external_id=ext, kind="cash", cash_value=val,
            ))
        elif rtype == "custody":
            val = parse_num(cell(r, i_val_base)) or parse_num(cell(r, i_val_acct))
            name = str(cell(r, i_name) or "").strip()
            if val is None or not name:
                continue
            ret_eur = parse_num(cell(r, i_ret_eur))
            purch = parse_num(cell(r, i_purch))
            ret_pct = None
            if purch and purch > 0:
                ret_pct = round((val - purch) / purch * 100, 2)
            elif ret_eur is not None and (val - ret_eur) != 0:
                ret_pct = round(ret_eur / (val - ret_eur) * 100, 2)
            key = acct_str or "Nordea custody"
            ext = "csv:nordea:" + re.sub(r"\s", "", key)[:40]
            if key not in invest_accts:
                invest_accts[key] = Account(
                    broker="Nordea", label=key, external_id=ext, kind="investment")
            invest_accts[key].holdings.append(Holding(
                name=name, value_eur=val,
                units=parse_num(cell(r, i_units)),
                isin=(str(cell(r, i_isin)).strip() if cell(r, i_isin) else None),
                return_pct=ret_pct, return_eur=ret_eur,
                currency=(str(cell(r, i_cur)).strip() if cell(r, i_cur) else None),
            ))
    accounts = list(invest_accts.values()) + cash_accts
    if not accounts:
        warnings.append(f"{filename}: no holdings or cash accounts parsed")
    return ParsedFile(source="nordea_xlsx", as_of=as_of, accounts=accounts, warnings=warnings)


# ── dispatch ─────────────────────────────────────────────────────────

def detect_and_parse(filename: str, raw: bytes) -> ParsedFile:
    low = (filename or "").lower()
    if low.endswith(".xlsx"):
        return parse_nordea_xlsx(filename, raw)
    if low.endswith(".csv"):
        return parse_nordnet_csv(filename, raw)
    raise ImportError_(f"{filename}: unsupported file type (need .csv or .xlsx)")


# ── Phase 1 verification CLI (read-only) ─────────────────────────────

def _verify():
    dl = os.path.expanduser("~/Downloads")
    candidates = [
        "Osaketaulukko salkkunro 18318444 24.5.2026.csv",
        "Rahastotaulukko salkkunro 18318444 24.5.2026.csv",
        "Omistukset.xlsx",
    ]
    print("Investment import — Phase 1 parse check (read-only)\n")
    grand = 0.0
    for fn in candidates:
        path = os.path.join(dl, fn)
        if not os.path.exists(path):
            print(f"– {fn}: not found, skipping")
            continue
        try:
            pf = detect_and_parse(fn, open(path, "rb").read())
        except Exception as e:
            print(f"✗ {fn}: {e}")
            continue
        print(f"● {fn}  [{pf.source}]  as_of={pf.as_of}")
        for a in pf.accounts:
            tag = "cash" if a.kind == "cash" else f"{len(a.holdings)} holdings"
            print(f"    {a.broker} ▸ {a.label}  ({tag})  total=€{a.total_eur:,.2f}")
            for h in a.holdings:
                dev = f"{h.return_pct:+.1f}%" if h.return_pct is not None else "—"
                pcs = f"{h.units:g}" if h.units is not None else "—"
                print(f"        · {h.name[:34]:<34} pcs={pcs:<10} €{h.value_eur:>10,.2f}  dev={dev}")
            grand += a.total_eur
        for w in pf.warnings:
            print("    ⚠", w)
        print()
    print(f"Grand total across files: €{grand:,.2f}  (read-only — nothing written)")
    return 0


if __name__ == "__main__":
    raise SystemExit(_verify())
